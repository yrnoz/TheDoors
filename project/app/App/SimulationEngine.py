from app.Database.ManageDB import *
from openpyxl import Workbook
import random
import time


SimRooms = db["Rooms"]  # create new table that called SimRooms just for the Simulation

# SimEmployees contains: id, permission, schedule, friends
SimEmployees = db["Employees"]  # create new table that called SimEmployees just for the Simulation


def simulation_import_room_details_from_file(input_file):
    """
    The function gets a CSV file with details about rooms in the
    factory and add them to the DB
    input: CSV file
    output: side effect  - the details added to the DB
    """
    global SimRooms
    with open(input_file) as details:  # open the file
        for line in filter(lambda x: x.strip(), details.readlines()):
            id, capacity, permission, floor = line[:-1].split(",")  # get the parameters we need from the line
            room = {"id": id, "capacity": int(capacity), "permission": int(permission), "floor": int(floor),
                    "schedule": {}}
            SimRooms.insert(room)  # add room's details to the DB



def simulation_export_rooms_to_file(output_file):
    """
    The function export to the manager a CSV file with all the rooms in the simulation factory
   :param output_file: file to write the rooms to.
   """


    global SimRooms
    with open(output_file, 'w') as output:
        for room in SimRooms.find():
            output.write(room["id"] + "," + str(room["capacity"]) + "," + str(room["permission"]) + ","
                         + str(room["floor"]) + "\n")


def simulation_add_random_employees(employees_number, min_permission):
    """
    The function add random employees to the simulation factory
    :param employees_number: number of employees in the simulation factory
    :param min_permission: min permission of employee.
    """
    global SimEmployees
    for i in range(0, employees_number):
        id = random.randint(0, employees_number * 1000)
        while check_id_of_employee(id):
            id = random.randint(0, employees_number * 1000)
        permission = random.randint(1, min_permission)
        employee = {"id": int(id), "name" : "John", "role" : "Engineer", "permission": int(permission), "password" : "123", "friends": [], "schedule": {}}
        Employees.insert(employee)  # add employee's details to the DB


def simulation_export_employees_to_file(output_file):
    global SimEmployees
    with open(output_file, 'w') as output:
        for employee in SimEmployees.find():
            output.write(str(employee["id"]) + "," + str(employee["permission"]) + \
                         simulation_get_schedule_of_employee(employee["id"]) + "\n")


def simulation_get_schedule_of_employee(id):
    employee = find_employee(id)
    schedule = employee["schedule"]
    employee_schedule = ""
    for date_time, tuple in schedule.items():
        employee_schedule += date_time + ": " + tuple[1] + "; "
    return employee_schedule

#the function copy all employees from the DB to the simulation DB
def simulation_copy_employees_from_DB():
    global SimEmployees
    for employee in Employees.objects():
        employee.schedules = ListField(Schedule)
        SimEmployees.insert(employee)

# the function copy all rooms from the DB to the simulation DB
def simulation_copy_rooms_from_DB():
    global SimRooms
    for room in Rooms.objects():
        room.schedules = ListField(Schedule)
        SimRooms.insert(room)

#the function assign employees with random people to rooms
def simulation_assign_employees(date_time, percent_employees):
    #now = time.strftime("%d/%m/%Y %H")
    rooms = SimRooms.objects()
    employees = SimEmployees.objects()
    count = 0
    for employee in employees:
        count += 1
        i = random.randint(0, len(rooms))
        num_employees = random.randint(0, int(len(employees)/4))
        while simulation_assign_employees_to_room_one_hour(date_time, rooms[i], num_employees, employee):
            i = random.randint(0, len(rooms))
        if count >= int((percent_employees*employees)/100):
            break



def simulation_assign_employees_to_room_one_hour(date_time, room, num_employees, employee):
    """
    this function gets date time in format "D/M/Y Hour", the room from the DB to assign employees,
    and number of employees to assign, if possible - they would be assigned to the room.
    :param employee:
    :param date_time: date time in format "D/M/Y Hour"
    :param room: the room from the DB to assign employees
    :param num_employees: number of employees to assign
    :return:  False - in case the employees can not be assigned to the room
              True - in case the employees were assigned to the room
    """
    global SimRooms
    capacity = room["capacity"]
    schedule = room["schedule"]
    schedule_employee = employee["schedule"]
    try:
        datetime.strptime(date_time, "%d/%m/%y %H")  # check the date_time format is correct
    except ValueError:
        return False
    if not (date_time in schedule):
        if num_employees > capacity or (date_time in schedule_employee):
            return False
        schedule[date_time] = (num_employees, None)
        schedule_employee[date_time] = (num_employees, room["id"])
        simulation_update_schedule_employees(date_time, room["id"], employee["id"], num_employees)
    else:
        if schedule[date_time][0] + num_employees > capacity | (date_time in schedule_employee):
            return False
        schedule[date_time] = (schedule[date_time][0] + num_employees, None)
        schedule_employee[date_time] = (num_employees, room["id"])
        simulation_update_schedule_employees(date_time, room["id"], employee["id"] ,num_employees)
    SimRooms.replace_one({'_id': room['_id']}, room)
    return True

def simulation_update_schedule_employees(date_time, room_id, employee_id, num_employees):
    schedule_employee = find_employee(employee_id)["schedule"]
    schedule_employee[date_time] = (num_employees, room_id)


#
#THIS IS THE MAIN FUNCTION OF THE SIMULATION
#
def simulation_day_in_factory(start_time, finish_time, percent_employees, new_rooms_details = None):
    simulation_copy_employees_from_DB()
    simulation_copy_rooms_from_DB()
    wb = Workbook()
    ws = wb.active()
    if not (new_rooms_details is None):
        simulation_import_room_details_from_file(new_rooms_details)
    date_now = time.strftime("%d/%m/%y")
    for hour in range(start_time, finish_time, 1):
        date_now += " " + str(hour)
        simulation_assign_employees(date_now, percent_employees)
    simulation_hist = dict()
    rooms = SimRooms.find()
    for hour in range(start_time, finish_time, 1):
        date_now += " " + str(hour)
        simulation_hist[date_now] = dict()
        i = 0
        for room in rooms:
            i+=1
            schedule = room["schedule"][date_now]
            simulation_hist[date_now].update({room["id"] : (schedule[0], room["capacity"])})
            ws["A"+str(i)] = room["id"]
            ws["B"+str(i)] = float(schedule[0])/int(room["capacity"])*100
    wb.save('Simulation_Stats.xlsx')












